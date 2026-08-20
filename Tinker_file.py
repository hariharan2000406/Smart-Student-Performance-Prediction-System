import tkinter as tk
from tkinter import messagebox
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── prediction logic ──────────────────────────────────────────────────────────
def predict(attendance, study_hours, internal_marks, assignment, prev_score):
    score = (
        attendance      * 0.25 +
        (study_hours / 10) * 100 * 0.15 +
        internal_marks  * 0.25 +
        assignment      * 0.15 +
        prev_score      * 0.20
    )

    if score >= 75:
        prediction = "Excellent Performance"
        risk       = "Low Risk"
        rec        = "Keep it up! Aim for distinction."
    elif score >= 60:
        prediction = "Good Performance"
        risk       = "Moderate Risk"
        rec        = "Focus on weak subjects and revise regularly."
    elif score >= 45:
        prediction = "Average Performance"
        risk       = "High Risk"
        rec        = "Attend extra classes and increase study hours."
    else:
        prediction = "Poor Performance"
        risk       = "Very High Risk"
        rec        = "Urgent: seek faculty help and improve attendance."

    return prediction, risk, rec

# ── main window ───────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("Smart Student Performance Prediction System")
root.geometry("700x560")
root.resizable(False, False)
root.configure(bg="white")

# ── input validation commands ─────────────────────────────────────────────────
def _val_id(P):
    """Student ID: digits only"""
    return P == "" or P.isdigit()

def _val_name(P):
    """Student Name: letters and spaces only"""
    return P == "" or all(c.isalpha() or c == " " for c in P)

def _val_float(P):
    """Attendance / Internal Marks / Assignment / Previous Score: float 0–100"""
    if P == "" or P == ".":
        return True
    try:
        v = float(P)
        return 0.0 <= v <= 100.0
    except ValueError:
        return False

def _val_hours(P):
    """Study Hours: float 0–8"""
    if P == "" or P == ".":
        return True
    try:
        v = float(P)
        return 0.0 <= v <= 8.0
    except ValueError:
        return False

vcmd_id    = (root.register(_val_id),    "%P")
vcmd_name  = (root.register(_val_name),  "%P")
vcmd_float = (root.register(_val_float), "%P")
vcmd_hours = (root.register(_val_hours), "%P")

# ── header frame ──────────────────────────────────────────────────────────────
header_frame = tk.Frame(root, bg="grey", pady=12)
header_frame.pack(fill="x")

tk.Label(
    header_frame,
    text="SMART STUDENT PERFORMANCE\nPREDICTION SYSTEM",
    font=("Arial", 16, "bold"),
    fg="white",
    bg="grey",
    justify="center"
).pack()

# ── content area ──────────────────────────────────────────────────────────────
content = tk.Frame(root, bg="white")
content.pack(fill="both", expand=True, padx=20, pady=10)

# ── student information frame ─────────────────────────────────────────────────
stu_frame = tk.LabelFrame(content, text="Student Information",
                          font=("Arial", 10, "bold"), bg="red",
                          fg="grey", padx=10, pady=10)
stu_frame.grid(row=0, column=0, padx=(0, 10), pady=5, sticky="nsew")

tk.Label(stu_frame, text="Student ID",   bg="white").grid(row=0, column=0, sticky="w", pady=4)
tk.Label(stu_frame, text="Student Name", bg="white").grid(row=1, column=0, sticky="w", pady=4)

entry_id   = tk.Entry(stu_frame, width=22, validate="key", validatecommand=vcmd_id)
entry_name = tk.Entry(stu_frame, width=22, validate="key", validatecommand=vcmd_name)
entry_id  .grid(row=0, column=1, padx=8)
entry_name.grid(row=1, column=1, padx=8)

# ── academic information frame ────────────────────────────────────────────────
acad_frame = tk.LabelFrame(content, text="Academic Information",
                           font=("Arial", 10, "bold"), bg="blue",
                           fg="blue", padx=10, pady=10)
acad_frame.grid(row=0, column=1, pady=5, sticky="nsew")

fields = [
    ("Attendance (%)",        "0–100"),
    ("Study Hours (per day)", "0–8"),
    ("Internal Marks (%)",    "0–100"),
    ("Assignment (%)",        "0–100"),
    ("Previous Score (%)",    "0–100"),
]

acad_entries = {}
# map each field label to its validation command
_field_vcmd = {
    "Attendance (%)":        vcmd_float,
    "Study Hours (per day)": vcmd_hours,
    "Internal Marks (%)":    vcmd_float,
    "Assignment (%)":        vcmd_float,
    "Previous Score (%)":    vcmd_float,
}
for i, (label, hint) in enumerate(fields):
    tk.Label(acad_frame, text=label, bg="white").grid(row=i, column=0, sticky="w", pady=3)
    e = tk.Entry(acad_frame, width=18, validate="key", validatecommand=_field_vcmd[label])
    e.grid(row=i, column=1, padx=8)
    acad_entries[label] = e

content.columnconfigure(0, weight=1)
content.columnconfigure(1, weight=1)

# ── action frame ──────────────────────────────────────────────────────────────
action_frame = tk.Frame(root, bg="#f0f2f5")
action_frame.pack(pady=6)

def on_predict():
    try:
        att   = float(acad_entries["Attendance (%)"].get())
        hrs   = float(acad_entries["Study Hours (per day)"].get())
        intm  = float(acad_entries["Internal Marks (%)"].get())
        asgn  = float(acad_entries["Assignment (%)"].get())
        prev  = float(acad_entries["Previous Score (%)"].get())
    except ValueError:
        messagebox.showerror("Input Error", "Please enter valid numbers in all academic fields.")
        return

    pred, risk, rec = predict(att, hrs, intm, asgn, prev)
    lbl_pred.config(text=pred)
    lbl_risk.config(text=risk,
                    fg="red" if "High" in risk else "#27ae60")
    txt_rec.config(state="normal")
    txt_rec.delete("1.0", "end")
    txt_rec.insert("1.0", rec)
    txt_rec.config(state="disabled")

def on_clear():
    entry_id.delete(0, "end")
    entry_name.delete(0, "end")
    for e in acad_entries.values():
        e.delete(0, "end")
    lbl_pred.config(text="")
    lbl_risk.config(text="", fg="grey")
    txt_rec.config(state="normal")
    txt_rec.delete("1.0", "end")
    txt_rec.config(state="disabled")

# ── save to excel ─────────────────────────────────────────────────────────────
EXCEL_FILE = "student_performance_log.xlsx"

HEADERS = [
    "Student ID", "Student Name",
    "Attendance (%)", "Study Hours (per day)", "Internal Marks (%)",
    "Assignment (%)", "Previous Score (%)",
    "Prediction", "Risk Level", "Recommendation"
]

def _header_style(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

def _get_or_create_wb():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Log"
        # write headers
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            _header_style(cell)
        # column widths
        widths = [14, 20, 16, 22, 18, 16, 18, 24, 16, 45]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
    return wb, ws

def on_save():
    # validate inputs exist
    stu_id   = entry_id.get().strip()
    stu_name = entry_name.get().strip()
    pred_text = lbl_pred.cget("text")

    if not stu_id or not stu_name:
        messagebox.showerror("Save Error", "Please enter Student ID and Student Name before saving.")
        return
    if not pred_text:
        messagebox.showerror("Save Error", "Please click 'Predict Performance' first before saving.")
        return

    try:
        att  = float(acad_entries["Attendance (%)"].get())
        hrs  = float(acad_entries["Study Hours (per day)"].get())
        intm = float(acad_entries["Internal Marks (%)"].get())
        asgn = float(acad_entries["Assignment (%)"].get())
        prev = float(acad_entries["Previous Score (%)"].get())
    except ValueError:
        messagebox.showerror("Save Error", "Academic fields have invalid values.")
        return

    risk = lbl_risk.cget("text")
    txt_rec.config(state="normal")
    rec  = txt_rec.get("1.0", "end").strip()
    txt_rec.config(state="disabled")

    wb, ws = _get_or_create_wb()
    next_row = ws.max_row + 1

    row_data = [stu_id, stu_name, att, hrs, intm, asgn, prev, pred_text, risk, rec]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # alternate row fill
    fill_color = "EBF3FB" if (next_row % 2 == 0) else "FFFFFF"
    row_fill = PatternFill("solid", fgColor=fill_color)

    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = border
        cell.fill      = row_fill

    ws.row_dimensions[next_row].height = 22

    wb.save(EXCEL_FILE)
    messagebox.showinfo("Saved", f"Record saved to '{EXCEL_FILE}'\nRow {next_row - 1} added.")

# ── buttons ───────────────────────────────────────────────────────────────────
tk.Button(action_frame, text="📊  Predict Performance",
          bg="blue", fg="white", font=("Arial", 10, "bold"),
          padx=12, pady=6, relief="flat", command=on_predict
          ).grid(row=0, column=0, padx=8)

tk.Button(action_frame, text="○  Clear",
          bg="orange", fg="white", font=("Arial", 10, "bold"),
          padx=12, pady=6, relief="flat", command=on_clear
          ).grid(row=0, column=1, padx=8)

tk.Button(action_frame, text="💾  Save to Excel",
          bg="#27ae60", fg="white", font=("Arial", 10, "bold"),
          padx=12, pady=6, relief="flat", command=on_save
          ).grid(row=0, column=2, padx=8)

tk.Button(action_frame, text="✕  Exit",
          bg="red", fg="white", font=("Arial", 10, "bold"),
          padx=12, pady=6, relief="flat", command=root.destroy
          ).grid(row=0, column=3, padx=8)

# ── result frame ──────────────────────────────────────────────────────────────
result_frame = tk.LabelFrame(root, text="Prediction Results",
                             font=("Arial", 10, "bold"), bg="green",
                             fg="grey", padx=14, pady=10)
result_frame.pack(fill="x", padx=20, pady=(4, 14))

tk.Label(result_frame, text="Prediction:",    bg="white", anchor="w").grid(row=0, column=0, sticky="w", pady=3)
tk.Label(result_frame, text="Risk Level:",    bg="white", anchor="w").grid(row=1, column=0, sticky="w", pady=3)
tk.Label(result_frame, text="Recommendation:",bg="white", anchor="w").grid(row=2, column=0, sticky="nw", pady=3)

lbl_pred = tk.Label(result_frame, text="", bg="white", font=("Arial", 10, "bold"), fg="#2c3e50")
lbl_risk = tk.Label(result_frame, text="", bg="white", font=("Arial", 10, "bold"))

lbl_pred.grid(row=0, column=1, sticky="w", padx=10)
lbl_risk.grid(row=1, column=1, sticky="w", padx=10)

txt_rec = tk.Text(result_frame, height=2, width=55, state="disabled",
                  bg="white", relief="flat", wrap="word")
txt_rec.grid(row=2, column=1, sticky="w", padx=10)

result_frame.columnconfigure(1, weight=1)

# ── run ───────────────────────────────────────────────────────────────────────
root.mainloop()
